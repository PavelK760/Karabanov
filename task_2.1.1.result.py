import csv
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side


class Vacancy:
    currencytorub = {
        "AZN": 35.68, "BYR": 23.91, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "EUR": 59.90, "UAH": 1.64, "UZS": 0.0055, "USD": 60.66,
    }

    def __init__(selfies, vacanciesy):
        selfies.name = vacanciesy['name']
        selfies.salary_from = int(float(vacanciesy['salary_from']))
        selfies.salary_to = int(float(vacanciesy['salary_to']))
        selfies.salary_currency = vacanciesy['salary_currency']
        selfies.salary_average = selfies.currencytorub[selfies.salary_currency] * (selfies.salary_from + selfies.salary_to) / 2
        selfies.area_name = vacanciesy['area_name']
        selfies.year = int(vacanciesy['published_at'][:4])



class DataSet:
    def __init__(selfies, names_file, names_vacancy):
        selfies.names_vacancy = names_vacancy
        selfies.names_file = names_file

    @staticmethod
    def aver(dict):
        new_dict = {}
        return DataSet.func1(dict, new_dict)

    @staticmethod
    def func1(dict, new_dict):
        return DataSet.func2(dict, new_dict)

    @staticmethod
    def func2(dict, new_dict):
        return DataSet.func3(dict, new_dict)

    @staticmethod
    def func3(dict, new_dict):
        return DataSet.func4(dict, new_dict)

    @staticmethod
    def func4(dict, new_dict):
        for keys, value in dict.items():
            new_dict[keys] = int(sum(value) / len(value))
        return new_dict

    @staticmethod
    def increments(dict, keys, amounts):
        DataSet.func5(amounts, dict, keys)

    @staticmethod
    def func5(amounts, dict, keys):
        if keys in dict:
            dict[keys] += amounts
        else:
            dict[keys] = amounts

    def csvreaders(selfies):
        yield from selfies.func6()

    def func6(selfies):
        with open(selfies.names_file, mode='r', encoding='utf-8-sig') as file:
            readers = csv.reader(file)
            headers = next(readers)
            headers_length = len(headers)
            yield from selfies.funca1(headers, headers_length, readers)

    def funca1(selfies, headers, headers_length, readers):
        for rows in readers:
            if '' not in rows and len(rows) == headers_length:
                yield dict(zip(headers, rows))

    def stats(selfies):
        salaries = {}
        city = {}
        salaryOfVacancy = {}
        count = 0

        count = selfies.funca2(city, count, salaries, salaryOfVacancy)

        number_vacancies = dict([(keys, len(values)) for keys, values in salaries.items()])
        namevac_number = dict([(keys, len(values)) for keys, values in salaryOfVacancy.items()])

        namevac_number, salaryOfVacancy = selfies.fucna3(namevac_number, number_vacancies, salaries, salaryOfVacancy)

        statist, statist2, statist3 = selfies.funca4(city, salaries, salaryOfVacancy)

        statist4, statist5 = selfies.funca5(city, count)
        statist3, statist5 = selfies.funca6(statist3, statist4, statist5)

        return statist, number_vacancies, statist2, namevac_number, statist3, statist5

    def funca6(selfies, statist3, statist4, statist5):
        statist4 = dict(statist4)
        statist3 = list(
            filter(lambda a: a[0] in list(statist4.keys()), [(key, value) for key, value in statist3.items()]))
        statist3.sort(key=lambda a: a[-1], reverse=True)
        statist3 = dict(statist3[:10])
        statist5 = dict(statist5[:10])
        return statist3, statist5

    def funca5(selfies, city, count):
        statist4 = {}
        selfies.func8(city, count, statist4)
        statist4 = list(filter(lambda a: a[-1] >= 0.01, [(keys, values) for keys, values in statist4.items()]))
        statist4.sort(key=lambda a: a[-1], reverse=True)
        statist5 = statist4.copy()
        return statist4, statist5

    def funca4(selfies, city, salaries, salaryOfVacancy):
        statist = selfies.aver(salaries)
        statist2 = selfies.aver(salaryOfVacancy)
        statist3 = selfies.aver(city)
        return statist, statist2, statist3

    def fucna3(selfies, namevac_number, number_vacancies, salaries, salaryOfVacancy):
        if not salaryOfVacancy:
            salaryOfVacancy = dict([(keys, [0]) for keys, values in salaries.items()])
            namevac_number = dict([(keys, 0) for keys, values in number_vacancies.items()])
        return namevac_number, salaryOfVacancy

    def funca2(selfies, city, count, salaries, salaryOfVacancy):
        for vacancies_dictionaries in selfies.csvreaders():
            vacanciesy = Vacancy(vacancies_dictionaries)
            selfies.increments(salaries, vacanciesy.year, [vacanciesy.salary_average])
            count = selfies.func7(city, count, salaryOfVacancy, vacanciesy)
        return count

    def func8(selfies, city, count, statist4):
        for years, salaries in city.items():
            statist4[years] = round(len(salaries) / count, 4)

    def func7(selfies, city, count, salaryOfVacancy, vacanciesy):
        if vacanciesy.name.find(selfies.names_vacancy) != -1:
            selfies.increments(salaryOfVacancy, vacanciesy.year, [vacanciesy.salary_average])
        selfies.increments(city, vacanciesy.area_name, [vacanciesy.salary_average])
        count += 1
        return count

    @staticmethod
    def printstats(statist1, statist2, statist3, statist4, statist5, statist6):
        print('Динамика уровня зарплат по годам: {0}'.format(statist1))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statist4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statist5))
        print('Динамика количества вакансий по годам: {0}'.format(statist2))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statist6))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statist3))


class InputConnect:
    def __init__(selfies):
        selfies.names_file = input('Введите название файла: ')
        selfies.names_vacancy = input('Введите название профессии: ')

        datasets = DataSet(selfies.names_file, selfies.names_vacancy)
        statist1, statist2, statist3, statist4, statist5, statist6 = datasets.stats()
        datasets.printstats(statist1, statist2, statist3, statist4, statist5, statist6)

        reported = Report(selfies.names_vacancy, statist1, statist2, statist3, statist4, statist5, statist6)
        reported.excel()

class Report:
    def __init__(selfies, names_vacanciesy, statist1, statist2, statist3, statist4, statist5, statist6):
        selfies.web = Workbook()
        selfies.names_vacancies = names_vacanciesy
        selfies.funca5(statist1, statist2, statist3, statist4, statist5, statist6)

    def funca5(selfies, statist1, statist2, statist3, statist4, statist5, statist6):
        selfies.statist1 = statist1
        selfies.statist2 = statist2
        selfies.statist3 = statist3
        selfies.statist4 = statist4
        selfies.statist5 = statist5
        selfies.statist6 = statist6

    def excel(selfies):
        wb1 = selfies.web.active
        wb1.title = 'Статистика по годам'
        wb1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + selfies.names_vacancies, 'Количество вакансий', 'Количество вакансий - ' + selfies.names_vacancies])
        selfies.funca6(wb1)

        datas = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + selfies.names_vacancies, ' Количество вакансий', ' Количество вакансий - ' + selfies.names_vacancies]]
        columnWidths = []
        columnWidths = selfies.func9(columnWidths, datas)

        selfies.func10(columnWidths, wb1)

        datas = []
        datas.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        selfies.func11(datas)
        ws2 = selfies.web.create_sheet('Статистика по городам')
        selfies.funca7(datas, ws2)

        columnWidths = []
        columnWidths = selfies.func12(columnWidths, datas)

        selfies.func14(columnWidths, ws2)

        fontBold = Font(bold=True)
        selfies.func13(fontBold, wb1, ws2)

        selfies.funca8(ws2)

        thin = Side(border_style='thin', color='00000000')

        selfies.funca9(datas, thin, ws2)

        selfies.statist1[1] = 1
        selfies.funca10(thin, wb1)

        selfies.web.save('report.xlsx')

    def funca10(selfies, thin, wb1):
        for rows, _ in enumerate(selfies.statist1):
            for column in 'ABCDE':
                wb1[column + str(rows + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def funca9(selfies, datas, thin, ws2):
        for rows in range(len(datas)):
            for column in 'ABDE':
                ws2[column + str(rows + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def funca8(selfies, ws2):
        for indexes, _ in enumerate(selfies.statist5):
            ws2['E' + str(indexes + 2)].number_format = '0.00%'

    def funca7(selfies, datas, ws2):
        for rows in datas:
            ws2.append(rows)

    def funca6(selfies, wb1):
        for years in selfies.statist1.keys():
            wb1.append([years, selfies.statist1[years], selfies.statist2[years], selfies.statist3[years],
                        selfies.statist4[years]])

    def func14(selfies, columnWidths, ws2):
        for j, columnWidth in enumerate(columnWidths, 1):
            ws2.column_dimensions[get_column_letter(j)].width = columnWidth + 2

    def func13(selfies, fontBold, wb1, ws2):
        for column in 'ABCDE':
            wb1[column + '1'].font = fontBold
            ws2[column + '1'].font = fontBold

    def func12(selfies, columnWidths, datas):
        for rows in datas:
            for j, cells in enumerate(rows):
                cells = str(cells)
                if len(columnWidths) > j:
                    if len(cells) > columnWidths[j]:
                        columnWidths[j] = len(cells)
                else:
                    columnWidths += [len(cells)]
        return columnWidths

    def func11(selfies, datas):
        for (city1, value1), (city2, value2) in zip(selfies.statist5.items(), selfies.statist6.items()):
            datas.append([city1, value1, '', city2, value2])

    def func10(selfies, columnWidths, wb1):
        selfies.func15(columnWidths, wb1)

    def func15(selfies, columnWidths, wb1):
        for j, columnWidth in enumerate(columnWidths, 1):
            wb1.column_dimensions[get_column_letter(j)].width = columnWidth + 2

    def func9(selfies, columnWidths, datas):
        for rows in datas:
            for j, cells in enumerate(rows):
                if len(columnWidths) > j:
                    if len(cells) > columnWidths[j]:
                        columnWidths[j] = len(cells)
                else:
                    columnWidths += [len(cells)]
        return columnWidths

if __name__ == '__main__':
    InputConnect()
