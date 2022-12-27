import csv
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import matplotlib.pyplot as plt
from openpyxl.styles import Font, Border, Side
import numpy as np

class Vacancy:
    currencytorub = {
        "AZN": 35.68, "BYR": 23.91, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "EUR": 59.90, "UAH": 1.64, "UZS": 0.0055, "USD": 60.66,
    }

    def __init__(selfies, vacanciesy):
        selfies.name = vacanciesy['name']
        selfies.salary_from = int(float(vacanciesy['salary_from']))
        selfies.salary_to = int(float(vacanciesy['salary_to']))
        selfies.salary_currency = vacanciesy['salary_currency']
        selfies.salary_average = selfies.currencytorub[selfies.salary_currency] * (selfies.salary_from + selfies.salary_to) / 2
        selfies.area_name = vacanciesy['area_name']
        selfies.year = int(vacanciesy['published_at'][:4])



class DataSet:
    def __init__(selfies, names_file, names_vacancy):
        selfies.names_vacancy = names_vacancy
        selfies.names_file = names_file

    @staticmethod
    def aver(dict):
        new_dict = {}
        return DataSet.func1(dict, new_dict)

    @staticmethod
    def func1(dict, new_dict):
        return DataSet.func2(dict, new_dict)

    @staticmethod
    def func2(dict, new_dict):
        return DataSet.func3(dict, new_dict)

    @staticmethod
    def func3(dict, new_dict):
        return DataSet.func4(dict, new_dict)

    @staticmethod
    def func4(dict, new_dict):
        for keys, value in dict.items():
            new_dict[keys] = int(sum(value) / len(value))
        return new_dict

    @staticmethod
    def increments(dict, keys, amounts):
        DataSet.func5(amounts, dict, keys)

    @staticmethod
    def func5(amounts, dict, keys):
        if keys in dict:
            dict[keys] += amounts
        else:
            dict[keys] = amounts

    def csvreaders(selfies):
        yield from selfies.func6()

    def func6(selfies):
        with open(selfies.names_file, mode='r', encoding='utf-8-sig') as file:
            readers = csv.reader(file)
            headers = next(readers)
            headers_length = len(headers)
            yield from selfies.funca1(headers, headers_length, readers)

    def funca1(selfies, headers, headers_length, readers):
        for rows in readers:
            if '' not in rows and len(rows) == headers_length:
                yield dict(zip(headers, rows))

    def stats(selfies):
        salaries = {}
        city = {}
        salaryOfVacancy = {}
        count = 0

        count = selfies.func9(city, count, salaries, salaryOfVacancy)

        number_vacancies = dict([(keys, len(values)) for keys, values in salaries.items()])
        namevac_number = dict([(keys, len(values)) for keys, values in salaryOfVacancy.items()])

        namevac_number, salaryOfVacancy = selfies.func10(namevac_number, number_vacancies, salaries, salaryOfVacancy)

        statist = selfies.aver(salaries)
        statist2 = selfies.aver(salaryOfVacancy)
        statist3 = selfies.aver(city)

        statist3, statist5 = selfies.func12(city, count, statist3)

        return statist, number_vacancies, statist2, namevac_number, statist3, statist5

    def func12(selfies, city, count, statist3):
        statist4, statist5 = selfies.func11(city, count)
        statist4 = dict(statist4)
        statist3 = list(
            filter(lambda a: a[0] in list(statist4.keys()), [(key, value) for key, value in statist3.items()]))
        statist3.sort(key=lambda a: a[-1], reverse=True)
        statist3 = dict(statist3[:10])
        statist5 = dict(statist5[:10])
        return statist3, statist5

    def func11(selfies, city, count):
        statist4 = {}
        selfies.func8(city, count, statist4)
        statist4 = list(filter(lambda a: a[-1] >= 0.01, [(keys, values) for keys, values in statist4.items()]))
        statist4.sort(key=lambda a: a[-1], reverse=True)
        statist5 = statist4.copy()
        return statist4, statist5

    def func10(selfies, namevac_number, number_vacancies, salaries, salaryOfVacancy):
        if not salaryOfVacancy:
            salaryOfVacancy = dict([(keys, [0]) for keys, values in salaries.items()])
            namevac_number = dict([(keys, 0) for keys, values in number_vacancies.items()])
        return namevac_number, salaryOfVacancy

    def func9(selfies, city, count, salaries, salaryOfVacancy):
        for vacancies_dictionaries in selfies.csvreaders():
            vacanciesy = Vacancy(vacancies_dictionaries)
            selfies.increments(salaries, vacanciesy.year, [vacanciesy.salary_average])
            count = selfies.func7(city, count, salaryOfVacancy, vacanciesy)
        return count

    def func8(selfies, city, count, statist4):
        for years, salaries in city.items():
            statist4[years] = round(len(salaries) / count, 4)

    def func7(selfies, city, count, salaryOfVacancy, vacanciesy):
        if vacanciesy.name.find(selfies.names_vacancy) != -1:
            selfies.increments(salaryOfVacancy, vacanciesy.year, [vacanciesy.salary_average])
        selfies.increments(city, vacanciesy.area_name, [vacanciesy.salary_average])
        count += 1
        return count

    @staticmethod
    def printstats(statist1, statist2, statist3, statist4, statist5, statist6):
        print('Динамика уровня зарплат по годам: {0}'.format(statist1))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statist4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statist5))
        print('Динамика количества вакансий по годам: {0}'.format(statist2))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statist6))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statist3))


class InputConnect:
    def __init__(selfies):
        selfies.names_file = input('Введите название файла: ')
        selfies.names_vacancies = input('Введите название профессии: ')
        # selfies.names_file = '../data/vacancies_by_year.csv'
        # selfies.names_vacancies = 'Программист'

        Setdata = DataSet(selfies.names_file, selfies.names_vacancies)
        statist1, statist2, statist3, statist4, statist5, statist6 = Setdata.stats()
        Setdata.printstats(statist1, statist2, statist3, statist4, statist5, statist6)
        statist1, statist2, statist3, statist4, statist5, statist6 = Setdata.stats()

        report = Report(selfies.names_vacancies, statist1, statist2, statist3, statist4, statist5, statist6)
        report.image()
        report.excel()
        report.save('report.xlsx')

class Report:
    def __init__(selfies, names_vacancies, statist1, statist2, statist3, statist4, statist5, statist6):
        selfies.wb = Workbook()
        selfies.names_vacancy = names_vacancies
        selfies.statist1 = statist1
        selfies.statist3 = statist3
        selfies.statist2 = statist2
        selfies.statist5 = statist5
        selfies.statist4 = statist4
        selfies.statist6 = statist6

    def excel(selfies):
        ws1 = selfies.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + selfies.names_vacancy, 'Количество вакансий', 'Количество вакансий - ' + selfies.names_vacancy])
        selfies.func13(ws1)

        datas = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + selfies.names_vacancy, ' Количество вакансий', ' Количество вакансий - ' + selfies.names_vacancy]]
        columnWth = []
        columnWth = selfies.func14(columnWth, datas)

        selfies.fucn15(columnWth, ws1)

        datas = []
        datas.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        selfies.func15(datas)
        ws2 = selfies.wb.create_sheet('Статистика по городам')
        for rows in datas:
            ws2.append(rows)

        columnWth = []
        columnWth = selfies.func16(columnWth, datas)

        selfies.fucn15(columnWth, ws2)

        fontBold = Font(bold=True)
        selfies.func18(fontBold, ws1, ws2)

        for indexes, _ in enumerate(selfies.statist5):
            ws2['E' + str(indexes + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        selfies.func19(datas, thin, ws2)

        selfies.func20(thin, ws1)

    def func20(selfies, thin, ws1):
        for rows, _ in enumerate(selfies.statist1):
            for column in 'ABCDE':
                ws1[column + str(rows + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def func19(selfies, datas, thin, ws2):
        for rows in range(len(datas)):
            for column in 'ABDE':
                ws2[column + str(rows + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def func18(selfies, fontBold, ws1, ws2):
        for column in 'ABCDE':
            ws1[column + '1'].font = fontBold
            ws2[column + '1'].font = fontBold

    def func16(selfies, columnWth, datas):
        for rows in datas:
            columnWth = selfies.func17(columnWth, rows)
        return columnWth

    def func17(selfies, columnWth, rows):
        for j, celles in enumerate(rows):
            celles = str(celles)
            if len(columnWth) > j:
                if len(celles) > columnWth[j]:
                    columnWth[j] = len(celles)
            else:
                columnWth += [len(celles)]
        return columnWth

    def func15(selfies, datas):
        for (city1, value1), (city2, value2) in zip(selfies.statist5.items(), selfies.statist6.items()):
            datas.append([city1, value1, '', city2, value2])

    def fucn15(selfies, columnWth, ws1):
        for j, columnWidth in enumerate(columnWth, 1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(j)].width = columnWidth + 2

    def func14(selfies, columnWth, datas):
        for rows in datas:
            for j, celles in enumerate(rows):
                if len(columnWth) > j:
                    if len(celles) > columnWth[j]:
                        columnWth[j] = len(celles)
                else:
                    columnWth += [len(celles)]
        return columnWth

    def func13(selfies, ws1):
        for years in selfies.statist1.keys():
            ws1.append([years, selfies.statist1[years], selfies.statist3[years], selfies.statist2[years],
                        selfies.statist4[years]])

    def image(selfies):
        fig, ((axel1, axel2), (axel3, axel4)) = plt.subplots(nrows=2, ncols=2)

        bariel1, bariel2 = selfies.func_bariel1(axel1)
        selfies.funct_axel1(axel1, bariel1, bariel2)

        axel2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bariel1, bariel2 = selfies.bariel12(axel2, bariel1, bariel2)
        selfies.funcaxel2(axel2, bariel1, bariel2)

        selfies.funcaxel3(axel3)

        selfies.funcaxel4(axel4)

        plt.tight_layout()
        plt.savefig('graph.png')

    def funcaxel4(selfies, axel4):
        axel4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in selfies.statist6.values()])
        axel4.pie(list(selfies.statist6.values()) + [other], labels=list(selfies.statist6.keys()) + ['Другие'],
                  textprops={'fontsize': 6})

    def funcaxel3(selfies, axel3):
        axel3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        axel3.barh(
            list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(selfies.statist5.keys()))]),
            list(reversed(list(selfies.statist5.values()))), color='blue', height=0.5, align='center')
        axel3.yaxis.set_tick_params(labelsize=6)
        axel3.xaxis.set_tick_params(labelsize=8)
        axel3.grid(axis='x')

    def funcaxel2(selfies, axel2, bariel1, bariel2):
        axel2.legend((bariel1[0], bariel2[0]),
                     ('Количество вакансий', 'Количество вакансий\n' + selfies.names_vacancy.lower()), prop={'size': 8})
        axel2.set_xticks(np.array(list(selfies.statist2.keys())) - 0.2, list(selfies.statist2.keys()), rotation=90)
        axel2.grid(axis='y')
        axel2.xaxis.set_tick_params(labelsize=8)
        axel2.yaxis.set_tick_params(labelsize=8)

    def bariel12(selfies, axel2, bariel1, bariel2):
        bariel1 = axel2.bar(np.array(list(selfies.statist2.keys())) - 0.4, selfies.statist2.values(), width=0.4)
        bariel2 = axel2.bar(np.array(list(selfies.statist2.keys())), selfies.statist4.values(), width=0.4)
        return bariel1, bariel2

    def funct_axel1(selfies, axel1, bariel1, bariel2):
        axel1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        axel1.grid(axis='y')
        axel1.legend((bariel1[0], bariel2[0]), ('средняя з/п', 'з/п ' + selfies.names_vacancy.lower()),
                     prop={'size': 8})
        axel1.set_xticks(np.array(list(selfies.statist1.keys())) - 0.2, list(selfies.statist1.keys()), rotation=90)
        axel1.xaxis.set_tick_params(labelsize=8)
        axel1.yaxis.set_tick_params(labelsize=8)

    def func_bariel1(selfies, axel1):
        bariel1 = axel1.bar(np.array(list(selfies.statist1.keys())) - 0.4, selfies.statist1.values(), width=0.4)
        bariel2 = axel1.bar(np.array(list(selfies.statist1.keys())), selfies.statist3.values(), width=0.4)
        return bariel1, bariel2

    def save(selfies, filename):
        selfies.wb.save(filename=filename)


if __name__ == '__main__':
    InputConnect()
